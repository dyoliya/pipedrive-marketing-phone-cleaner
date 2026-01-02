# -------------------------ABOUT --------------------------

# pyinstaller --onefile tool_ui.py
# Tool: Pipedrive Marketing Cleaning Tool
# Developer: dyoliya
# Created: 2025-08-06

# © 2025 dyoliya. All rights reserved.

# ---------------------------------------------------------


import os
import re
from glob import glob
from datetime import datetime
import pandas as pd
from tqdm import tqdm
from io import StringIO
from io import BytesIO
from collections import defaultdict
from config.gdrive_client import download_file_by_id, list_files_in_folder
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import json

with open("config/gdrive_files.json", "r") as f:
    GDRIVE_FILES = json.load(f)

with open("config/gdrive_folders.json", "r") as f:
    GDRIVE_FOLDERS = json.load(f)

# ----------------------- DIRECTORIES -----------------------
# input folder
INPUT_FOLDER = "for_processing"
os.makedirs("for_processing", exist_ok=True)
os.makedirs(INPUT_FOLDER, exist_ok=True)

# output folders
OUTPUT_CLEANED_FOLDER = "output"
os.makedirs("output", exist_ok=True)
os.makedirs(OUTPUT_CLEANED_FOLDER, exist_ok=True)

# ----------------------- PHONE FIELDS -----------------------

PHONE_FIELDS = [
    "Person - Phone - Work",
    "Person - Phone - Home",
    "Person - Phone - Mobile",
    "Person - Phone - Other",
    "Person - Phone 1",
    "Person - Phone 2",
    "Person - Phone 3",
    "Person - Phone 4",
    "Person - Phone 5",
    "Person - Phone 6",
    "Person - Phone 7",
    "Person - Phone 8",
    "Person - Phone 9",
    "Person - Phone 10",
    "Person - Archive - Phone"

]

# ------------------ FUNCTIONS ------------------
def check_required_columns(df, file_path):
    required_columns = [
        "Deal - ID",
        "Deal - Contact person",
        "Deal - Owner",
        "Deal - County",
        "Deal - Stage"
    ]

    missing = [col for col in required_columns if col not in df.columns]
    has_phone_field = any(col in df.columns for col in PHONE_FIELDS)

    if missing or not has_phone_field:
        msg_parts = []
        if missing:
            msg_parts.append(f"\n- {', '.join(missing)}")
        if not has_phone_field:
            msg_parts.append("\n- at least 1 valid phone field")
        
        print(f"\n⚠️  Skipping {os.path.basename(file_path)} due to missing column/s:{''.join(msg_parts)}")
        return False

    return True

def normalize_phone(number):
    return re.sub(r"[^\d]", "", str(number))

def load_opt_out_phone_numbers(excel_filenames=None):
    # Default Excel files if none specified
    if excel_filenames is None:
        excel_filenames = ["DNC (Cold-PD).xlsx", "CallTextOut-7d (PD).xlsx"]
    
    numbers = defaultdict(set)  # phone -> set of filenames

    for name in excel_filenames:
        clean_name = name.replace(".xlsx", "")
        file_id = GDRIVE_FILES.get(clean_name)

        if not file_id:
            print(f"⚠️ Missing GDrive file ID for {name}")
            continue

        try:
            content = download_file_by_id(file_id)
            xls = pd.ExcelFile(content)

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=str)
                if df.empty or 0 not in df.columns:
                    continue

                nums = df[0].dropna().astype(str).map(normalize_phone)
                for num in nums:
                    numbers[num].add(name)

        except Exception as e:
            print(f"Error reading GDrive file {name}: {e}")
    return numbers

def load_pd_phone_numbers():
    pd_phone_numbers = {}
    folder_id = GDRIVE_FOLDERS["pd_phone"]

    try:
        files = list_files_in_folder(folder_id)

        for file in files:
            if not file["name"].endswith(".xlsx"):
                continue

            try:
                content = download_file_by_id(file["id"])
                df = pd.read_excel(content, engine="openpyxl", dtype=str)
                df.fillna("", inplace=True)

                for _, row in df.iterrows():
                    deal_id = row.get("Deal - ID", "")
                    deal_stage = row.get("Deal - Stage", "")

                    for field in PHONE_FIELDS:
                        raw_phones = str(row.get(field, ""))
                        if not raw_phones.strip():
                            continue

                        for phone in map(str.strip, raw_phones.split(",")):
                            normalized = normalize_phone(phone)
                            if len(normalized) == 11 and normalized.startswith("1"):
                                normalized = normalized[1:]
                            if len(normalized) == 10 and normalized.isdigit():
                                # Store as list to handle multiple deals
                                pd_phone_numbers.setdefault(normalized, []).append({
                                    "deal_id": deal_id,
                                    "deal_stage": deal_stage
                                })

            except Exception as e:
                print(f"Error reading file {file['name']}: {e}")

    except Exception as e:
        print(f"Error reading GDrive pd_phone folder: {e}")

    return pd_phone_numbers

def extract_first_name(contact_person, deal_title):
    name = str(contact_person).strip()

    # Normalize: remove all extra spaces (including between letters) and lowercase
    normalized = re.sub(r"[^a-zA-Z]", "", name).lower()

    # Check for placeholders like "noname" or "unknown"
    if not name or normalized in ["noname", "unknown", "uunknown", "nunknown"]:
        title = str(deal_title).strip()
        if title and not re.match(r"(?i)^no name|^unknown", title):
            return title.split()[0].capitalize()
        return ""

    # Otherwise, take first word before space or slash
    first_word = re.split(r"[ /]", name)[0].strip()

    return first_word.capitalize()

def extract_deal_owner(deal_owner):
    if pd.isna(deal_owner):
        return ""
    owner = str(deal_owner).strip()
    return owner.split()[0] if owner else ""

def format_deal_county(deal_county):
    if pd.isna(deal_county) or not str(deal_county).strip():
        return ""

    counties = [c.strip() for c in str(deal_county).split(",") if c.strip()]
    grouped = [", ".join(counties[i:i+2]) for i in range(0, len(counties), 2)]

    if len(grouped) == 1:
        return f'"{grouped[0]}"'
    elif len(grouped) == 2:
        return f'"{grouped[0]} and {grouped[1]}"'
    else:
        return f'"{", ".join(grouped[:-1])} and {grouped[-1]}"'


# ------------------ MAIN SCRIPT ------------------

def main():
    seen_normalized_numbers = {}
    pd_phone_numbers = load_pd_phone_numbers()
    input_files = glob(os.path.join(INPUT_FOLDER, "*.xlsx"))
    cleaned_data = {}
    
    opt_out_cache = {}

    for file_path in tqdm(input_files, desc="Processing input files"):
        try:
            df = pd.read_excel(file_path, engine='openpyxl', dtype=str)
            if not check_required_columns(df, file_path):
                continue
            df.fillna("", inplace=True)
            cleaned_rows = []
            

            for idx, row in tqdm(df.iterrows(), total=len(df), desc=f"Processing {os.path.basename(file_path)}", leave=False):
                try:
                    remarks = ""
                    deal_id = row.get("Deal - ID", "")
                    deal_stage = row.get("Deal - Stage", "")
                    if deal_stage == "Cold Deals - Priority 2":
                        key = "cold"
                        excel_files_to_check = [
                            "DNC (Cold-PD).xlsx",
                            "CallOut-14d+TextOut-30d (Cold).xlsx"
                        ]
                    else:
                        key = "normal"
                        excel_files_to_check = [
                            "DNC (Cold-PD).xlsx",
                            "CallTextOut-7d (PD).xlsx"
                        ]

                    if key not in opt_out_cache:
                        opt_out_cache[key] = load_opt_out_phone_numbers(excel_files_to_check)
                        print(excel_files_to_check)

                    gdrive_numbers = opt_out_cache[key]

                    contact_person = row.get("Deal - Contact person", "")
                    deal_title = row.get("Deal - Title", "")
                    first_name = extract_first_name(contact_person, deal_title)
                    deal_owner = row.get("Deal - Owner", "")
                    deal_owner_fn = extract_deal_owner(deal_owner)
                    raw_county = row.get("Deal - County", "")
                    formatted_county = format_deal_county(raw_county)

                    duplicates_found = {}
                    phone_to_use = ""
                    disallowed_found = False

                    opt_out_remarks = []
                    pd_phone_remarks = []
                    duplicate_remarks = []
                    format_remarks = []

                    # ------------------ STEP 1: OPT-OUT CHECK FIRST ------------------
                    opt_out_matches = defaultdict(list)
                    remaining_numbers = []

                    for field in PHONE_FIELDS:
                        raw_phones = str(row.get(field, "")).strip()
                        if not raw_phones:
                            continue
                        for phone in map(str.strip, raw_phones.split(",")):
                            normalized = normalize_phone(phone)
                            if len(normalized) == 11 and normalized.startswith("1"):
                                normalized = normalized[1:]
                            if not (len(normalized) == 10 and normalized.isdigit()):
                                format_remarks.append(
                                    f"Phone number {phone} has incorrect format even after normalization"
                                )
                                continue

                            if normalized in gdrive_numbers:
                                for fname in gdrive_numbers[normalized]:
                                    opt_out_matches[fname].append(normalized)
                            else:
                                remaining_numbers.append(normalized)

                    # Add opt-out remarks if any phones found in opt-out
                    if opt_out_matches:
                        # parts = []
                        for fname, nums in opt_out_matches.items():
                            plural = "numbers" if len(nums) > 1 else "number"
                            opt_out_remarks.append(f"Phone {plural} {', '.join(nums)} exist in {fname}")
                        remarks = "; ".join(opt_out_remarks)
                        # Don't set disallowed_found = True yet, because some phones remain


                    # ------------------ STEP 2: PD PHONE CHECK ------------------
                    if not disallowed_found and remaining_numbers:
                        pd_phone_remarks = set()

                        for normalized in remaining_numbers:
                            if normalized in pd_phone_numbers:
                                existing_entries = pd_phone_numbers[normalized]
                                for entry in existing_entries:
                                    existing_deal = entry["deal_id"]
                                    existing_stage = entry["deal_stage"]
                                    current_stage = row.get("Deal - Stage", "")

                                    # Only block if stage is different
                                    if existing_stage != current_stage:
                                        pd_phone_remarks.add(
                                            f"{normalized} exists in Deal ID {existing_deal} on stage {existing_stage} (PD Phone Numbers)"
                                        )
                                        disallowed_found = True

                        # Combine remarks
                        if pd_phone_remarks:
                            if remarks:
                                remarks += "; " + "; ".join(pd_phone_remarks)
                            else:
                                remarks = "; ".join(pd_phone_remarks)


                        # If still no disallow, keep first unique number
                        if not disallowed_found:
                            for normalized in remaining_numbers:
                                if normalized in seen_normalized_numbers:
                                    first_deal = seen_normalized_numbers[normalized]
                                    if first_deal != deal_id:
                                        duplicates_found[normalized] = first_deal
                                        duplicate_remarks.append(
                                            f"Phone number {normalized} already exists in Deal ID {first_deal}"
                                        )
                                else:
                                    seen_normalized_numbers[normalized] = deal_id

                                    if not phone_to_use:
                                        phone_to_use = normalized


                    # ------------------ STEP 3: Remarks & Fallbacks ------------------                    
                    remarks_list = []
                    if format_remarks:
                        remarks_list.append("; ".join(format_remarks))
                    if opt_out_remarks:
                        remarks_list.append("; ".join(opt_out_remarks))
                    if pd_phone_remarks:
                        remarks_list.append("; ".join(pd_phone_remarks))
                    if duplicate_remarks:
                        remarks_list.append("; ".join(duplicate_remarks))

                    remarks = "; ".join(remarks_list)

                    # ----- STEP 4: Final Cleaning to Retain Numbers with No Remarks -----
                    non_formatting_remarks = [r for r in remarks_list if r not in format_remarks]
                    # Apply final rules
                    if not phone_to_use:
                        # Case 1: No valid phone, keep all remarks
                        remarks = "; ".join(remarks_list)  
                    elif phone_to_use and not non_formatting_remarks:
                        # Case 2: Valid phone exists and remarks only about formatting
                        remarks = ""  # discard formatting-only remarks
                    elif phone_to_use and non_formatting_remarks:
                        # Case 4: Valid phone exists, but there are other critical issues (opt-out / duplicate)
                        phone_to_use = ""  # remove all phones
                        remarks = "; ".join(remarks_list)  # keep all remarks
                    else:
                        # Fallback (no remarks, no phone)
                        remarks = ""

                    # ------------------ STEP 5: Append to Cleaned Rows ------------------
                    cleaned_rows.append({
                            "Carrier": "",
                            "Deal - ID": deal_id,
                            "Phone Number": phone_to_use,
                            "First Name": first_name,
                            "Deal - Value": row.get("Deal - Value", ""),
                            "Deal - Owner": deal_owner_fn,
                            "Deal - County": formatted_county,
                            "Deal - Title": row.get("Deal - Title", ""),
                            "Deal - Stage": deal_stage,
                            "Remarks": remarks
                    })

                except Exception as row_err:
                    print(f"⚠️ Skipping row {idx} in {os.path.basename(file_path)}: {row_err}")


            if cleaned_rows:
                cleaned_df = pd.DataFrame(cleaned_rows)
                sheet_name = os.path.splitext(os.path.basename(file_path))[0][:31]
                cleaned_data[sheet_name] = cleaned_df

        except Exception as e:
            print(f"Error processing {file_path}: {e}")

 # ------------- COMBINE INTO ONE EXCEL FILE -------------
    if cleaned_data:
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        combined_output = os.path.join(OUTPUT_CLEANED_FOLDER, f"{date_str}_pd_mktg_combined_output.xlsx")

        with pd.ExcelWriter(combined_output, engine="openpyxl") as writer:
            for sheet_name, df in cleaned_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Add empty 'carrier' sheet
            empty_df = pd.DataFrame()
            empty_df.to_excel(writer, sheet_name="carrier", index=False)

        # ------------- ADD FORMULAS USING OPENPYXL -------------
        wb = load_workbook(combined_output)
        for sheet_name in cleaned_data.keys():
            ws = wb[sheet_name]

            # Find column index for "Carrier" and "Phone Number"
            headers = [cell.value for cell in ws[1]]
            try:
                carrier_col = headers.index("Carrier") + 1
                phone_col = headers.index("Phone Number") + 1
            except ValueError:
                continue  # skip if missing

            # Apply formula =VLOOKUP(C2,carrier!A:C,3,FALSE)
            for row_idx in range(2, ws.max_row + 1):
                formula = f'=VLOOKUP({get_column_letter(phone_col)}{row_idx},carrier!A:C,3,FALSE)'
                ws.cell(row=row_idx, column=carrier_col).value = formula

        wb.save(combined_output)
        print(f"\n✅ Combined cleaned file saved to: {combined_output}")
    
if __name__ == "__main__":
    main()